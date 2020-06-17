from tkinter import * # импортируем весь tkinter
import openpyxl       #импортируем для тривьев, таблицы
import os,sys         #импортируем библиотеку что бы открыть файл

#функция для просмотра технологических карт         
        
def cl(selection2):
    if sheet1['C2'].value == selection2:
        root1=Tk()
        root1.title(sheet1['C2'].value)
        lab=Label(root1,text=sheet1['C1'].value)
        lab.grid()
    else:
        root2=Tk()
        root2.title(sheet2['C2'].value)
        lab=Label(root2,text=sheet2['C1'].value)
        lab.grid()
    
def balance():
    bal=Label(root,text,font="Arial 18",fg="white",bg="blue")
    bal.grid()

def callback(selection):
    global selproduct
    selproduct = selection

    
    
#функция для обработки склада(остатка)
def clicked():
    
    #res = y.get()  # взять значение из окошечка
    
    #messagebox.showinfo("GUI Python", message.get())
    cell = sheet3['A1']
    for row in sheet3.rows:
        print(cell.row)
        if row == selproduct: 
            #string = ''
            #string = string + str(cell.value) + ' '
            #print(string)
            print(cell.value)
                #for cell in row:     добавить фор для добавления значения в другую колонну
                #string = string + str(cell.value) + ' '
            sheet3['B1'].value = sheet3['B1'].value + int(y.get())
            wc.save('/home/dayugov/программа_для_Алёны/ostatok.xlsx')
        else:
            print("false")
            
    
        
        
        
        
    
def file():
    #открываем табличку Остаток
    os.system('chgrp -R /home/dayugov/программа_для_Алёны/ostatok.xlsx')


# создание экрана    
root=Tk()
root.title("моя программа версия 0.1")
root.geometry("1280x1080")

#переменные
selection = str()
selection2 = str()
selproduct = str()
#читаем excel-файл
wa = openpyxl.load_workbook('/home/dayugov/программа_для_Алёны/Biskvit_vanilniy_i_shokoladniy.xlsx')
wb = openpyxl.load_workbook('/home/dayugov/программа_для_Алёны/Marmelad_chernika.xlsx')
wc = openpyxl.load_workbook('/home/dayugov/программа_для_Алёны/ostatok.xlsx')

#активный лист
sheet1 = wa.active
sheet2 = wb.active
sheet3 = wc.active

#метка
lab=Label(root) #,text="",font="Arial 18", fg="white",bg="white")
lab.grid() # без него не отразится на экране
#кнопкиok
but1=Button(root,text="Приход") 
# width-ширина height-высота bg-цвет фона fg-цвет надписи
but1.grid()
but2=Button(root,text="Расход")
but2.grid()
but3=Button(root,text="Остаток",command = file)
but3.grid()

but4=StringVar(root)
but4.set("Технологические карты")  # default value
#список технологических карт, названия взяты из табличек
kor=[sheet1['C2'].value, sheet2['C2'].value]
#but4=OptionMenu(root, but4, "Бисквит Ванильный и шоколадный", "Мармелад черника") 
but4=OptionMenu(root, but4, *kor, command = cl)
# width-ширина height-высота bg-цвет фона fg-цвет надписи
but4.grid()

but5=Button(root,text="Дв.Тов.") 
# width-ширина height-высота bg-цвет фона fg-цвет надписи
but5.grid()

but6=StringVar(root)
but6.set("Выберите продукт и введите количество") # default value
kor2=[ "Пюре черники", "Сахар", "Глюкоза", "Желтый пектин(яблочный)", "Лимонная кислота", "Вода", "Мука пшеничная", "Ванильный экстракт", "Разрыхлитель", "Сметана 20%", "Сливочное масло 82,5%", "Яйца", "Соль", "Растительное масло", "Какао порошок Какао-барри (Брют)", "Сода"]
but6 = OptionMenu(root, but6, *kor2, command = callback)
but6.grid(column=2, row=3)

#message = int()
y = Entry(root, width = 15, justify = RIGHT)#, textvariable = message) 
y.grid(column=3, row=3)

but8=Button(root, text="Добавить", command = clicked)
but8.grid(column=4, row=3) 

but9=Button(root,text="Убрать")
but9.grid(column=5, row=3) 

root.mainloop()