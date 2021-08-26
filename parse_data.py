import os
from bs4 import BeautifulSoup
import argparse
import openpyxl
from openpyxl.styles import Font, Alignment
import pathlib
import sys
import string
import re


def get_file_list(directory):
    """
    Get list of files of specified directory

    Keyword arguments:
    directory -- destination directory
    """
    lst = []
    for file in os.listdir(directory):
        if os.path.isfile(os.path.join(directory, file)):
            lst.append(str(file))
    return lst


def re_wrap(field, text):
    match = re.search(r'(?<={}: ).*(?=\n)'.format(field), text)
    return match if match is None else match.group()


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--files', '-f', help="path to directory with downloaded files", type=str, default='output')
    parser.add_argument('--output', '-o', help="result file name", type=str, default='winetitles_media_producers')
    args = parser.parse_args()

    if not pathlib.Path(args.files).is_dir():
        sys.exit('path to files is incorrect')

    files = get_file_list(args.files)

    info = []
    hyperlinks = []
    skipped_files = []

    for file_name in files:
        file = open('{}/{}'.format(args.files, file_name), encoding="utf-8")
        soup = BeautifulSoup(file, 'html.parser')

        details_elem = soup.find('div', {'class': 'wid-full-contact-details'})
        if details_elem is None:
            skipped_files.append(file_name)
            continue
        row_elem = details_elem.findChild('div', {'class': 'row'})
        parsed = row_elem.text

        winery_name = soup.find('h3', {'class': 'post-title'}).text
        company_name = re_wrap('Company Name', row_elem.text)
        phone = re_wrap('Telephone', row_elem.text)
        fax = re_wrap('Fax', row_elem.text)
        address = re_wrap('Address', row_elem.text)
        website = re_wrap('Website', row_elem.text)
        brands = re_wrap('Brands', row_elem.text)

        info.append({
            'winery name': winery_name,
            'company name': company_name,
            'phone': phone,
            'fax': fax,
            'address': address,
            'website': website,
            'brands': brands
        })

        hyperlinks.append({
            'website': 'http://{}'.format(website) if website is not None else None
        })

        print(info[len(info) - 1])

        file.close()

    print('Skipped files: {}'.format(skipped_files))

    # Create xlsx
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()

    # List of used headers
    headers = ['winery name', 'company name', 'phone', 'fax', 'address', 'website', 'brands']

    # Make headers look pretty
    for i in range(len(headers)):
        cell = sheet.cell(1, i + 1)
        cell.value = headers[i]
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Used later to calculate width of the column
    max_length = {}
    for header in headers:
        max_length[header] = -1

    # Insert data
    for row in range(2, len(info) + 2):
        for i in range(len(headers)):
            # Insert value
            cell = sheet.cell(row, i + 1)

            # Insert hyperlink if column has one
            if headers[i] in hyperlinks[row - 2] and hyperlinks[row - 2][headers[i]] is not None:
                cell.style = "Hyperlink"
                cell.value = '=HYPERLINK("{}","{}")'.format(hyperlinks[row - 2][headers[i]],
                                                            info[row - 2][headers[i]])
            else:
                cell.value = info[row - 2][headers[i]]

            # Calculate max length
            if cell.value is not None:
                max_length[headers[i]] = max(max_length[headers[i]], len(cell.value))

    for i in range(0, len(headers)):
        sheet.column_dimensions[string.ascii_uppercase[i]].width = max_length[headers[i]] + 5

    wb.save(filename='{}.xlsx'.format(args.output))